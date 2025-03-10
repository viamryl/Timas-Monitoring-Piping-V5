import pandas as pd
from openpyxl import load_workbook
import datetime
import time
import shutil
import os
import logging
import warnings
import numpy as np
import traceback

def load_data(file_path, sheetname,skip_row_n, used_cols_list, add_pk = False):
    df = pd.read_excel(file_path,
                       sheet_name=sheetname,
                       skiprows=skip_row_n,
                       usecols=used_cols_list)
    
    if add_pk == True:
        df["PK"] = df["AREA"].astype(str) + df["LINE NO"].astype(str) + " " + df["MAT'L LINE"].astype(str) + " " + df["JOINT NO"].astype(str)
    
    return df

def number_to_excel_column(n):
    column = ""
    while n > 0:
        n -= 1  # Mengurangi 1 untuk menangani indeks berbasis 1
        column = chr(n % 26 + 65) + column  # 65 adalah ASCII 'A'
        n //= 26
    return column

def update_data(df1, df2, pk):
    diff = pd.merge(df1,
                    df2,
                    on = pk,
                    how = 'outer',
                    indicator=True)
    
    missing_data = diff[diff['_merge'] != 'both'][[pk, 'LINE NO_x', 'JOINT NO_x','LINE NO_y', "JOINT NO_y", '_merge']].reset_index(drop=True)
    missing_data['Missing From'] = missing_data['_merge'].map({'left_only': 'Added', 'right_only': 'Deleted'})
    missing_data = missing_data.drop(columns=['_merge'])
    missing_data["DATE"]=datetime.datetime.now().strftime('%Y/%m/%d')
    missing_data.reset_index(drop=True,inplace=True)
    missing_data.rename(columns={
                            "LINE NO_x" : "ADDED LINE NO",
                            "JOINT NO_x" : "ADDED JOINT NO",
                            "LINE NO_y" : "DELETED LINE NO",
                            "JOINT NO_y" : "DELETED JOINT NO",
                            "Missing From" : "STATUS"
                        },inplace = True)

    return missing_data

def write_data(df, destination_path, sheetname, startrow, startcol, pk = "PK", typedf = None):

    if typedf == "matl" or typedf == "mir" or typedf == "progress" or typedf == "claim":
        a = df.columns.tolist().index('DIA-INCH PLAN SW')+1
        b = df.columns.tolist().index('DIA-INCH PLAN FW')+1
        c = df.columns.tolist().index('TOTAL')+1
        aa = df.columns.tolist().index('FW / SW')+1
        bb = df.columns.tolist().index('DN')+1
        va = number_to_excel_column(a)
        vb = number_to_excel_column(b)
        vc = number_to_excel_column(c)
        vaa = number_to_excel_column(aa)
        vbb = number_to_excel_column(bb)
    if typedf == "progress" or typedf == "claim":
        d = df.columns.tolist().index('FIT-UP RECORD SW')+1
        e = df.columns.tolist().index('FIT-UP RECORD FW')+1
        f = df.columns.tolist().index('WELDING RECORD SW')+1
        g = df.columns.tolist().index('WELDING RECORD FW')+1
        h = df.columns.tolist().index('FIT-UP RECORD DATE')+1
        m = df.columns.tolist().index('WELDING RECORD DATE')+1
        vh = number_to_excel_column(h)
        vm = number_to_excel_column(m)
    if typedf == "claim":
        afivis = df.columns.tolist().index('QAQC VISUAL DATE')+1
        n = df.columns.tolist().index('PPC CLAIM REPORT FABS 30%')+1
        o = df.columns.tolist().index('PPC CLAIM REPORT INSTALL 60%')+1
        p = df.columns.tolist().index('PPC CLAIM REPORT PUNCHLIST 10%')+1
        q = df.columns.tolist().index('PPC CLAIM REPORT TOTAL')+1
        r = df.columns.tolist().index('CUMM')+1
        vafivis = number_to_excel_column(afivis)
        vn = number_to_excel_column(n)
        vo = number_to_excel_column(o)
        vp = number_to_excel_column(p)
        vq = number_to_excel_column(q)
    if typedf == "matl" or typedf == "mir":
        dba = number_to_excel_column(dbexcols.index('MATL CODE')+1)
        dbb = number_to_excel_column(dbexcols.index('TYPE')+1)
        dbc = number_to_excel_column(dbexcols.index('MATL TYPE')+1)
        dbd = number_to_excel_column(dbexcols.index('GROUP MATL')+1)
        dbe = number_to_excel_column(dbexcols.index('CLASS / SCH')+1)
        dbf = number_to_excel_column(dbexcols.index('BASE SIZE 1')+1)
        dbg = number_to_excel_column(dbexcols.index('BASE SIZE 2')+1)
        dbh = number_to_excel_column(dbexcols.index('THK 1')+1)
        dbi = number_to_excel_column(dbexcols.index('Standard Matl Name')+1)
        dbj = number_to_excel_column(dbexcols.index('UOM')+1)
    if typedf == "matl":
        pcode = df.columns.tolist().index('P MATL CODE')+1
        pa = df.columns.tolist().index('P TYPE')+1
        pb = df.columns.tolist().index('P MATL TYPE')+1
        pc = df.columns.tolist().index('P GROUP MATL')+1
        pd = df.columns.tolist().index('P CLASS / SCH')+1
        pe = df.columns.tolist().index('P BASE SIZE 1')+1
        pf = df.columns.tolist().index('P BASE SIZE 2')+1
        pg = df.columns.tolist().index('P THK 1')+1
        ph = df.columns.tolist().index('P Standard Matl Name')+1
        pi = df.columns.tolist().index('P UOM')+1

        scode = df.columns.tolist().index('S MATL CODE')+1
        sa = df.columns.tolist().index('S TYPE')+1
        sb = df.columns.tolist().index('S MATL TYPE')+1
        sc = df.columns.tolist().index('S GROUP MATL')+1
        sd = df.columns.tolist().index('S CLASS / SCH')+1
        se = df.columns.tolist().index('S BASE SIZE 1')+1
        sf = df.columns.tolist().index('S BASE SIZE 2')+1
        sg = df.columns.tolist().index('S THK 1')+1
        sh = df.columns.tolist().index('S Standard Matl Name')+1
        si = df.columns.tolist().index('S UOM')+1

        tcode = df.columns.tolist().index('T MATL CODE')+1
        ta = df.columns.tolist().index('T TYPE')+1
        tb = df.columns.tolist().index('T MATL TYPE')+1
        tc = df.columns.tolist().index('T GROUP MATL')+1
        td = df.columns.tolist().index('T CLASS / SCH')+1
        te = df.columns.tolist().index('T BASE SIZE 1')+1
        tf = df.columns.tolist().index('T BASE SIZE 2')+1
        tg = df.columns.tolist().index('T THK 1')+1
        th = df.columns.tolist().index('T Standard Matl Name')+1
        ti = df.columns.tolist().index('T UOM')+1

        vpcode = number_to_excel_column(pcode)
        vscode = number_to_excel_column(scode)
        vtcode = number_to_excel_column(tcode)
    if typedf == "mir":
        psrcode = df.columns.tolist().index('P SR / MIR NO')+1
        psra = df.columns.tolist().index('P SR DATE')+1
        psrb = df.columns.tolist().index('P PIC')+1
        psrc = df.columns.tolist().index('P QTY TOTAL')+1
        psrd = df.columns.tolist().index('P QTY HAULING')+1
        psre = df.columns.tolist().index('P QTY OUTSTANDING')+1
        psrf = df.columns.tolist().index('P QTY MAPPING')+1
        psrg = df.columns.tolist().index('P MAPPING BALANCE')+1
        psrph = df.columns.tolist().index('P Standard Matl Name')+1
        psrqty = df.columns.tolist().index('P QTY MATL')+1

        ssrcode = df.columns.tolist().index('S SR / MIR NO')+1
        ssra = df.columns.tolist().index('S SR DATE')+1
        ssrb = df.columns.tolist().index('S PIC')+1
        ssrc = df.columns.tolist().index('S QTY TOTAL')+1
        ssrd = df.columns.tolist().index('S QTY HAULING')+1
        ssre = df.columns.tolist().index('S QTY OUTSTANDING')+1
        ssrf = df.columns.tolist().index('S QTY MAPPING')+1
        ssrg = df.columns.tolist().index('S MAPPING BALANCE')+1
        ssrph = df.columns.tolist().index('S Standard Matl Name')+1
        ssrqty = df.columns.tolist().index('S QTY MATL')+1

        tsrcode = df.columns.tolist().index('T SR / MIR NO')+1
        tsra = df.columns.tolist().index('T SR DATE')+1
        tsrb = df.columns.tolist().index('T PIC')+1
        tsrc = df.columns.tolist().index('T QTY TOTAL')+1
        tsrd = df.columns.tolist().index('T QTY HAULING')+1
        tsre = df.columns.tolist().index('T QTY OUTSTANDING')+1
        tsrf = df.columns.tolist().index('T QTY MAPPING')+1
        tsrg = df.columns.tolist().index('T MAPPING BALANCE')+1
        tsrph = df.columns.tolist().index('T Standard Matl Name')+1
        tsrqty = df.columns.tolist().index('T QTY MATL')+1

        vpsrcode = number_to_excel_column(psrcode)
        vpsrf = number_to_excel_column(psrf)
        vpsrph = number_to_excel_column(psrph)
        vpsrqty = number_to_excel_column(psrqty)

        vssrcode = number_to_excel_column(ssrcode)
        vssrf = number_to_excel_column(ssrf)
        vssrph = number_to_excel_column(ssrph)
        vssrqty = number_to_excel_column(ssrqty)

        vtsrcode = number_to_excel_column(tsrcode)
        vtsrf = number_to_excel_column(tsrf)
        vtsrph = number_to_excel_column(tsrph)
        vtsrqty = number_to_excel_column(tsrqty)

        master_mir = number_to_excel_column(mirmastercols.index('SR / MIR NO')+1)
        master_matlname = number_to_excel_column(mirmastercols.index('Standard Matl Name')+1)
        master_date = number_to_excel_column(mirmastercols.index('SR DATE')+1)
        master_pic = number_to_excel_column(mirmastercols.index('PIC')+1)
        master_qty = number_to_excel_column(mirmastercols.index('QTY TOTAL')+1)
        master_hauling = number_to_excel_column(mirmastercols.index('QTY HAULING')+1)
        master_outstanding = number_to_excel_column(mirmastercols.index('QTY OUTSTANDING')+1)

    wb = load_workbook(destination_path)
    ws = wb[sheetname]

    if sheetname == "Monitoring Piping" :
        try:
            # ws.delete_rows(startrow, ws.max_row - startrow + 1)
            df.dropna(subset = pk, inplace = True)
            df = df[df[pk] != ""]
            df = df[df[pk] != "nan"]
            df = df[~df[pk].astype(str).str.startswith('nan')]
            df.reset_index(drop=True, inplace=True)
        except:
            pass
        for i, row in df.iterrows():
            for j, value in enumerate(row):
                ws.cell(row=startrow + i, column=startcol + j, value=value)

            if df["FW / SW"][i] == "SW" or df["FW / SW"][i] == "FW":
                ws.cell(row=startrow + i, column=a, value = f'=IF({vaa}{startrow + i}="SW",1,IF({vaa}{startrow + i}="FW",0,))*IF({vbb}{startrow + i}=15,0.5,IF({vbb}{startrow + i}=20,0.75,IF({vbb}{startrow + i}=25,1,IF({vbb}{startrow + i}=32,1.25,IF({vbb}{startrow + i}=40,1.5,IF({vbb}{startrow + i}=50,2,IF({vbb}{startrow + i}=65,2.5,IF({vbb}{startrow + i}=80,3,IF({vbb}{startrow + i}>=100,{vbb}{startrow + i}/25)))))))))')
                ws.cell(row=startrow + i, column=b, value = f'=IF({vaa}{startrow + i}="SW",0,IF({vaa}{startrow + i}="FW",1,))*IF({vbb}{startrow + i}=15,0.5,IF({vbb}{startrow + i}=20,0.75,IF({vbb}{startrow + i}=25,1,IF({vbb}{startrow + i}=32,1.25,IF({vbb}{startrow + i}=40,1.5,IF({vbb}{startrow + i}=50,2,IF({vbb}{startrow + i}=65,2.5,IF({vbb}{startrow + i}=80,3,IF({vbb}{startrow + i}>=100,{vbb}{startrow + i}/25)))))))))')
            ws.cell(row=startrow + i, column=c, value = f"={va}{startrow+i}+{vb}{startrow+i}")
            if typedf == "progress" or typedf == "claim":
                ws.cell(row=startrow + i, column=d, value = f'=IF({vh}{startrow + i}>0,{va}{startrow + i},"")')
                ws.cell(row=startrow + i, column=e, value = f'=IF({vh}{startrow + i}>0,{vb}{startrow + i},"")')
                ws.cell(row=startrow + i, column=f, value = f'=IF({vm}{startrow + i}>0,{va}{startrow + i},"")')
                ws.cell(row=startrow + i, column=g, value = f'=IF({vm}{startrow + i}>0,{vb}{startrow + i},"")')
            if typedf == "claim":
                ws.cell(row=startrow + i, column=n, value = f'=IF({vafivis}{startrow+i}>0,{vc}{startrow + i}*30%,"")')
                ws.cell(row=startrow + i, column=o, value = f'=IF({vafivis}{startrow+i}>0,{vc}{startrow + i}*60%,"")')
                ws.cell(row=startrow + i, column=p, value = f'=IF({vafivis}{startrow+i}>0,{vc}{startrow + i}*10%,"")')
                ws.cell(row=startrow + i, column=q, value = f'={vn}{startrow + i} + {vo}{startrow + i} + {vp}{startrow + i}')
                ws.cell(row=startrow + i, column=r, value = f'={vq}{startrow + i}')
            if typedf == "matl":
                ws.cell(row=startrow + i, column=pa, value = f"=IFERROR(IF(INDEX('DB MATL'!${dbb}:${dbb}, MATCH(${vpcode}{startrow + i}, 'DB MATL'!${dba}:${dba}, 0)) = 0, \"\", INDEX('DB MATL'!${dbb}:${dbb}, MATCH(${vpcode}{startrow + i}, 'DB MATL'!${dba}:${dba}, 0))), \"\")")
                ws.cell(row=startrow + i, column=pb, value = f"=IFERROR(IF(INDEX('DB MATL'!${dbc}:${dbc}, MATCH(${vpcode}{startrow + i}, 'DB MATL'!${dba}:${dba}, 0)) = 0, \"\", INDEX('DB MATL'!${dbc}:${dbc}, MATCH(${vpcode}{startrow + i}, 'DB MATL'!${dba}:${dba}, 0))), \"\")")
                ws.cell(row=startrow + i, column=pc, value = f"=IFERROR(IF(INDEX('DB MATL'!${dbd}:${dbd}, MATCH(${vpcode}{startrow + i}, 'DB MATL'!${dba}:${dba}, 0)) = 0, \"\", INDEX('DB MATL'!${dbd}:${dbd}, MATCH(${vpcode}{startrow + i}, 'DB MATL'!${dba}:${dba}, 0))), \"\")")
                ws.cell(row=startrow + i, column=pd, value = f"=IFERROR(IF(INDEX('DB MATL'!${dbe}:${dbe}, MATCH(${vpcode}{startrow + i}, 'DB MATL'!${dba}:${dba}, 0)) = 0, \"\", INDEX('DB MATL'!${dbe}:${dbe}, MATCH(${vpcode}{startrow + i}, 'DB MATL'!${dba}:${dba}, 0))), \"\")")
                ws.cell(row=startrow + i, column=pe, value = f"=IFERROR(IF(INDEX('DB MATL'!${dbf}:${dbf}, MATCH(${vpcode}{startrow + i}, 'DB MATL'!${dba}:${dba}, 0)) = 0, \"\", INDEX('DB MATL'!${dbf}:${dbf}, MATCH(${vpcode}{startrow + i}, 'DB MATL'!${dba}:${dba}, 0))), \"\")")
                ws.cell(row=startrow + i, column=pf, value = f"=IFERROR(IF(INDEX('DB MATL'!${dbg}:${dbg}, MATCH(${vpcode}{startrow + i}, 'DB MATL'!${dba}:${dba}, 0)) = 0, \"\", INDEX('DB MATL'!${dbg}:${dbg}, MATCH(${vpcode}{startrow + i}, 'DB MATL'!${dba}:${dba}, 0))), \"\")")
                ws.cell(row=startrow + i, column=pg, value = f"=IFERROR(IF(INDEX('DB MATL'!${dbh}:${dbh}, MATCH(${vpcode}{startrow + i}, 'DB MATL'!${dba}:${dba}, 0)) = 0, \"\", INDEX('DB MATL'!${dbh}:${dbh}, MATCH(${vpcode}{startrow + i}, 'DB MATL'!${dba}:${dba}, 0))), \"\")")
                ws.cell(row=startrow + i, column=ph, value = f"=IFERROR(IF(INDEX('DB MATL'!${dbi}:${dbi}, MATCH(${vpcode}{startrow + i}, 'DB MATL'!${dba}:${dba}, 0)) = 0, \"\", INDEX('DB MATL'!${dbi}:${dbi}, MATCH(${vpcode}{startrow + i}, 'DB MATL'!${dba}:${dba}, 0))), \"\")")
                ws.cell(row=startrow + i, column=pi, value = f"=IFERROR(IF(INDEX('DB MATL'!${dbj}:${dbj}, MATCH(${vpcode}{startrow + i}, 'DB MATL'!${dba}:${dba}, 0)) = 0, \"\", INDEX('DB MATL'!${dbj}:${dbj}, MATCH(${vpcode}{startrow + i}, 'DB MATL'!${dba}:${dba}, 0))), \"\")")
            
                ws.cell(row=startrow + i, column=sa, value = f"=IFERROR(IF(INDEX('DB MATL'!${dbb}:${dbb}, MATCH(${vscode}{startrow + i}, 'DB MATL'!${dba}:${dba}, 0)) = 0, \"\", INDEX('DB MATL'!${dbb}:${dbb}, MATCH(${vscode}{startrow + i}, 'DB MATL'!${dba}:${dba}, 0))), \"\")")
                ws.cell(row=startrow + i, column=sb, value = f"=IFERROR(IF(INDEX('DB MATL'!${dbc}:${dbc}, MATCH(${vscode}{startrow + i}, 'DB MATL'!${dba}:${dba}, 0)) = 0, \"\", INDEX('DB MATL'!${dbc}:${dbc}, MATCH(${vscode}{startrow + i}, 'DB MATL'!${dba}:${dba}, 0))), \"\")")
                ws.cell(row=startrow + i, column=sc, value = f"=IFERROR(IF(INDEX('DB MATL'!${dbd}:${dbd}, MATCH(${vscode}{startrow + i}, 'DB MATL'!${dba}:${dba}, 0)) = 0, \"\", INDEX('DB MATL'!${dbd}:${dbd}, MATCH(${vscode}{startrow + i}, 'DB MATL'!${dba}:${dba}, 0))), \"\")")
                ws.cell(row=startrow + i, column=sd, value = f"=IFERROR(IF(INDEX('DB MATL'!${dbe}:${dbe}, MATCH(${vscode}{startrow + i}, 'DB MATL'!${dba}:${dba}, 0)) = 0, \"\", INDEX('DB MATL'!${dbe}:${dbe}, MATCH(${vscode}{startrow + i}, 'DB MATL'!${dba}:${dba}, 0))), \"\")")
                ws.cell(row=startrow + i, column=se, value = f"=IFERROR(IF(INDEX('DB MATL'!${dbf}:${dbf}, MATCH(${vscode}{startrow + i}, 'DB MATL'!${dba}:${dba}, 0)) = 0, \"\", INDEX('DB MATL'!${dbf}:${dbf}, MATCH(${vscode}{startrow + i}, 'DB MATL'!${dba}:${dba}, 0))), \"\")")
                ws.cell(row=startrow + i, column=sf, value = f"=IFERROR(IF(INDEX('DB MATL'!${dbg}:${dbg}, MATCH(${vscode}{startrow + i}, 'DB MATL'!${dba}:${dba}, 0)) = 0, \"\", INDEX('DB MATL'!${dbg}:${dbg}, MATCH(${vscode}{startrow + i}, 'DB MATL'!${dba}:${dba}, 0))), \"\")")
                ws.cell(row=startrow + i, column=sg, value = f"=IFERROR(IF(INDEX('DB MATL'!${dbh}:${dbh}, MATCH(${vscode}{startrow + i}, 'DB MATL'!${dba}:${dba}, 0)) = 0, \"\", INDEX('DB MATL'!${dbh}:${dbh}, MATCH(${vscode}{startrow + i}, 'DB MATL'!${dba}:${dba}, 0))), \"\")")
                ws.cell(row=startrow + i, column=sh, value = f"=IFERROR(IF(INDEX('DB MATL'!${dbi}:${dbi}, MATCH(${vscode}{startrow + i}, 'DB MATL'!${dba}:${dba}, 0)) = 0, \"\", INDEX('DB MATL'!${dbi}:${dbi}, MATCH(${vscode}{startrow + i}, 'DB MATL'!${dba}:${dba}, 0))), \"\")")
                ws.cell(row=startrow + i, column=si, value = f"=IFERROR(IF(INDEX('DB MATL'!${dbj}:${dbj}, MATCH(${vscode}{startrow + i}, 'DB MATL'!${dba}:${dba}, 0)) = 0, \"\", INDEX('DB MATL'!${dbj}:${dbj}, MATCH(${vscode}{startrow + i}, 'DB MATL'!${dba}:${dba}, 0))), \"\")")            
            
                ws.cell(row=startrow + i, column=ta, value = f"=IFERROR(IF(INDEX('DB MATL'!${dbb}:${dbb}, MATCH(${vtcode}{startrow + i}, 'DB MATL'!${dba}:${dba}, 0)) = 0, \"\", INDEX('DB MATL'!${dbb}:${dbb}, MATCH(${vtcode}{startrow + i}, 'DB MATL'!${dba}:${dba}, 0))), \"\")")
                ws.cell(row=startrow + i, column=tb, value = f"=IFERROR(IF(INDEX('DB MATL'!${dbc}:${dbc}, MATCH(${vtcode}{startrow + i}, 'DB MATL'!${dba}:${dba}, 0)) = 0, \"\", INDEX('DB MATL'!${dbc}:${dbc}, MATCH(${vtcode}{startrow + i}, 'DB MATL'!${dba}:${dba}, 0))), \"\")")
                ws.cell(row=startrow + i, column=tc, value = f"=IFERROR(IF(INDEX('DB MATL'!${dbd}:${dbd}, MATCH(${vtcode}{startrow + i}, 'DB MATL'!${dba}:${dba}, 0)) = 0, \"\", INDEX('DB MATL'!${dbd}:${dbd}, MATCH(${vtcode}{startrow + i}, 'DB MATL'!${dba}:${dba}, 0))), \"\")")
                ws.cell(row=startrow + i, column=td, value = f"=IFERROR(IF(INDEX('DB MATL'!${dbe}:${dbe}, MATCH(${vtcode}{startrow + i}, 'DB MATL'!${dba}:${dba}, 0)) = 0, \"\", INDEX('DB MATL'!${dbe}:${dbe}, MATCH(${vtcode}{startrow + i}, 'DB MATL'!${dba}:${dba}, 0))), \"\")")
                ws.cell(row=startrow + i, column=te, value = f"=IFERROR(IF(INDEX('DB MATL'!${dbf}:${dbf}, MATCH(${vtcode}{startrow + i}, 'DB MATL'!${dba}:${dba}, 0)) = 0, \"\", INDEX('DB MATL'!${dbf}:${dbf}, MATCH(${vtcode}{startrow + i}, 'DB MATL'!${dba}:${dba}, 0))), \"\")")
                ws.cell(row=startrow + i, column=tf, value = f"=IFERROR(IF(INDEX('DB MATL'!${dbg}:${dbg}, MATCH(${vtcode}{startrow + i}, 'DB MATL'!${dba}:${dba}, 0)) = 0, \"\", INDEX('DB MATL'!${dbg}:${dbg}, MATCH(${vtcode}{startrow + i}, 'DB MATL'!${dba}:${dba}, 0))), \"\")")
                ws.cell(row=startrow + i, column=tg, value = f"=IFERROR(IF(INDEX('DB MATL'!${dbh}:${dbh}, MATCH(${vtcode}{startrow + i}, 'DB MATL'!${dba}:${dba}, 0)) = 0, \"\", INDEX('DB MATL'!${dbh}:${dbh}, MATCH(${vtcode}{startrow + i}, 'DB MATL'!${dba}:${dba}, 0))), \"\")")
                ws.cell(row=startrow + i, column=th, value = f"=IFERROR(IF(INDEX('DB MATL'!${dbi}:${dbi}, MATCH(${vtcode}{startrow + i}, 'DB MATL'!${dba}:${dba}, 0)) = 0, \"\", INDEX('DB MATL'!${dbi}:${dbi}, MATCH(${vtcode}{startrow + i}, 'DB MATL'!${dba}:${dba}, 0))), \"\")")
                ws.cell(row=startrow + i, column=ti, value = f"=IFERROR(IF(INDEX('DB MATL'!${dbj}:${dbj}, MATCH(${vtcode}{startrow + i}, 'DB MATL'!${dba}:${dba}, 0)) = 0, \"\", INDEX('DB MATL'!${dbj}:${dbj}, MATCH(${vtcode}{startrow + i}, 'DB MATL'!${dba}:${dba}, 0))), \"\")")
            if typedf == "mir":
                ws.cell(row=startrow + i, column=psra, value = f"=IF(OR({vpsrcode}{startrow + i}=\"\", {vpsrph}{startrow + i}=\"\"), \"\", IFERROR(LOOKUP(1, 1/(MASTER!{master_mir}:{master_mir}={vpsrcode}{startrow + i})/(MASTER!{master_matlname}:{master_matlname}={vpsrph}{startrow + i}), MASTER!{master_date}:{master_date}), \"\"))")
                ws.cell(row=startrow + i, column=psrb, value = f"=IF(OR({vpsrcode}{startrow + i}=\"\", {vpsrph}{startrow + i}=\"\"), \"\", IFERROR(LOOKUP(1, 1/(MASTER!{master_mir}:{master_mir}={vpsrcode}{startrow + i})/(MASTER!{master_matlname}:{master_matlname}={vpsrph}{startrow + i}), MASTER!{master_pic}:{master_pic}), \"\"))")
                ws.cell(row=startrow + i, column=psrc, value = f"=IF(OR({vpsrcode}{startrow + i}=\"\", {vpsrph}{startrow + i}=\"\"), \"\", IFERROR(LOOKUP(1, 1/(MASTER!{master_mir}:{master_mir}={vpsrcode}{startrow + i})/(MASTER!{master_matlname}:{master_matlname}={vpsrph}{startrow + i}), MASTER!{master_qty}:{master_qty}), \"\"))")
                ws.cell(row=startrow + i, column=psrd, value = f"=IF(OR({vpsrcode}{startrow + i}=\"\", {vpsrph}{startrow + i}=\"\"), \"\", IFERROR(LOOKUP(1, 1/(MASTER!{master_mir}:{master_mir}={vpsrcode}{startrow + i})/(MASTER!{master_matlname}:{master_matlname}={vpsrph}{startrow + i}), MASTER!{master_hauling}:{master_hauling}), \"\"))")
                ws.cell(row=startrow + i, column=psre, value = f"=IF(OR({vpsrcode}{startrow + i}=\"\", {vpsrph}{startrow + i}=\"\"), \"\", IFERROR(LOOKUP(1, 1/(MASTER!{master_mir}:{master_mir}={vpsrcode}{startrow + i})/(MASTER!{master_matlname}:{master_matlname}={vpsrph}{startrow + i}), MASTER!{master_outstanding}:{master_outstanding}), \"\"))")
                ws.cell(row=startrow + i, column=psrg, value = f"=IF(AND({vpsrf}{startrow + i}=\"\", {vpsrqty}{startrow + i}=\"\"),\"\", {vpsrf}{startrow + i} - {vpsrqty}{startrow + i})")

                ws.cell(row=startrow + i, column=ssra, value = f"=IF(OR({vssrcode}{startrow + i}=\"\", {vssrph}{startrow + i}=\"\"), \"\", IFERROR(LOOKUP(1, 1/(MASTER!{master_mir}:{master_mir}={vssrcode}{startrow + i})/(MASTER!{master_matlname}:{master_matlname}={vssrph}{startrow + i}), MASTER!{master_date}:{master_date}), \"\"))")
                ws.cell(row=startrow + i, column=ssrb, value = f"=IF(OR({vssrcode}{startrow + i}=\"\", {vssrph}{startrow + i}=\"\"), \"\", IFERROR(LOOKUP(1, 1/(MASTER!{master_mir}:{master_mir}={vssrcode}{startrow + i})/(MASTER!{master_matlname}:{master_matlname}={vssrph}{startrow + i}), MASTER!{master_pic}:{master_pic}), \"\"))")
                ws.cell(row=startrow + i, column=ssrc, value = f"=IF(OR({vssrcode}{startrow + i}=\"\", {vssrph}{startrow + i}=\"\"), \"\", IFERROR(LOOKUP(1, 1/(MASTER!{master_mir}:{master_mir}={vssrcode}{startrow + i})/(MASTER!{master_matlname}:{master_matlname}={vssrph}{startrow + i}), MASTER!{master_qty}:{master_qty}), \"\"))")
                ws.cell(row=startrow + i, column=ssrd, value = f"=IF(OR({vssrcode}{startrow + i}=\"\", {vssrph}{startrow + i}=\"\"), \"\", IFERROR(LOOKUP(1, 1/(MASTER!{master_mir}:{master_mir}={vssrcode}{startrow + i})/(MASTER!{master_matlname}:{master_matlname}={vssrph}{startrow + i}), MASTER!{master_hauling}:{master_hauling}), \"\"))")
                ws.cell(row=startrow + i, column=ssre, value = f"=IF(OR({vssrcode}{startrow + i}=\"\", {vssrph}{startrow + i}=\"\"), \"\", IFERROR(LOOKUP(1, 1/(MASTER!{master_mir}:{master_mir}={vssrcode}{startrow + i})/(MASTER!{master_matlname}:{master_matlname}={vssrph}{startrow + i}), MASTER!{master_outstanding}:{master_outstanding}), \"\"))")
                ws.cell(row=startrow + i, column=ssrg, value = f"=IF(AND({vssrf}{startrow + i}=\"\", {vssrqty}{startrow + i}=\"\"),\"\", {vssrf}{startrow + i} - {vssrqty}{startrow + i})")
              
                ws.cell(row=startrow + i, column=tsra, value = f"=IF(OR({vtsrcode}{startrow + i}=\"\", {vtsrph}{startrow + i}=\"\"), \"\", IFERROR(LOOKUP(1, 1/(MASTER!{master_mir}:{master_mir}={vtsrcode}{startrow + i})/(MASTER!{master_matlname}:{master_matlname}={vtsrph}{startrow + i}), MASTER!{master_date}:{master_date}), \"\"))")
                ws.cell(row=startrow + i, column=tsrb, value = f"=IF(OR({vtsrcode}{startrow + i}=\"\", {vtsrph}{startrow + i}=\"\"), \"\", IFERROR(LOOKUP(1, 1/(MASTER!{master_mir}:{master_mir}={vtsrcode}{startrow + i})/(MASTER!{master_matlname}:{master_matlname}={vtsrph}{startrow + i}), MASTER!{master_pic}:{master_pic}), \"\"))")
                ws.cell(row=startrow + i, column=tsrc, value = f"=IF(OR({vtsrcode}{startrow + i}=\"\", {vtsrph}{startrow + i}=\"\"), \"\", IFERROR(LOOKUP(1, 1/(MASTER!{master_mir}:{master_mir}={vtsrcode}{startrow + i})/(MASTER!{master_matlname}:{master_matlname}={vtsrph}{startrow + i}), MASTER!{master_qty}:{master_qty}), \"\"))")
                ws.cell(row=startrow + i, column=tsrd, value = f"=IF(OR({vtsrcode}{startrow + i}=\"\", {vtsrph}{startrow + i}=\"\"), \"\", IFERROR(LOOKUP(1, 1/(MASTER!{master_mir}:{master_mir}={vtsrcode}{startrow + i})/(MASTER!{master_matlname}:{master_matlname}={vtsrph}{startrow + i}), MASTER!{master_hauling}:{master_hauling}), \"\"))")
                ws.cell(row=startrow + i, column=tsre, value = f"=IF(OR({vtsrcode}{startrow + i}=\"\", {vtsrph}{startrow + i}=\"\"), \"\", IFERROR(LOOKUP(1, 1/(MASTER!{master_mir}:{master_mir}={vtsrcode}{startrow + i})/(MASTER!{master_matlname}:{master_matlname}={vtsrph}{startrow + i}), MASTER!{master_outstanding}:{master_outstanding}), \"\"))")
                ws.cell(row=startrow + i, column=tsrg, value = f"=IF(AND({vtsrf}{startrow + i}=\"\", {vtsrqty}{startrow + i}=\"\"),\"\", {vtsrf}{startrow + i} - {vtsrqty}{startrow + i})")

        total_rows_in_ws = ws.max_row
        rows_to_keep = startrow + len(df) - 1
        if total_rows_in_ws > rows_to_keep:
            ws.delete_rows(rows_to_keep + 1, total_rows_in_ws - rows_to_keep)


    elif sheetname == "Updates from Engineer": 
        total_rows_in_ws = ws.max_row
        df.reset_index(drop = True, inplace = True)
        for i, row in df.iterrows():
            for j, value in enumerate(row):
                ws.cell(row=1 + i + total_rows_in_ws, column=startcol + j, value=value)

    elif sheetname == "Conflict":
        total_rows_in_ws = ws.max_row
        df.reset_index(drop = True, inplace = True)
        for i, row, in df.iterrows():
            for j, value in enumerate(row):
                ws.cell(row = total_rows_in_ws + 1 + i, column=startcol+j, value=value)

    else :
        print("ngaco lu")
        print("yamaap")

    wb.save(destination_path)

def syncronized(df1, df2, pk1, pk2, how = 'left'):
    df1.drop_duplicates(subset=[pk1], inplace = True, keep = 'first')
    df1.dropna(subset=[pk1], inplace =True)

    df2.drop_duplicates(subset=[pk2], inplace = True, keep = 'first')
    df2.dropna(subset=[pk2], inplace =True)
    
    merged_data = pd.merge(df1,
                            df2,
                            left_on=pk1,
                            right_on=pk2,
                            how = how,
                            sort = False)
    
    merged_data[pk2] = merged_data[pk2].fillna(merged_data[pk1])

    return merged_data 

def production(df,dashboardpath,filename,dropdup = None,dropn = None, instance = False):
    if instance == False:
        df.drop_duplicates(subset=dropdup,keep = "first", inplace = True)
        df.dropna(subset = dropn, inplace=True)
        df = df[~df[dropdup].astype(str).str.startswith('nan')]
        df.to_csv(f"{dashboardpath}/{filename}.csv", index=False)
    else : 
        df.to_csv(f"{dashboardpath}/{filename}.csv", index=False)

def backup(destination_path, backup_path,filename):
    shutil.copy(destination_path,backup_path)
    os.rename(f"{backup_path}/{filename}.xlsx",
            f"{backup_path}/{filename} {datetime.datetime.now().strftime('%Y-%m-%d_%H-%M-%S-%f')[:-3]}.xlsx")
    
    status = "Data Has Been Writen Succesfully"
    return status

def slicing(path):
    slashindex = path.rfind('/')
    flname_ext = path[slashindex + 1:]
    last_dot_index = flname_ext.rfind('.')
    flname = flname_ext[:last_dot_index]

    return flname

def fixdate(df,cols):
    for col in cols:
        df[col] = pd.to_datetime(df[col], format = "mixed")
        df[col] = df[col].dt.strftime("%Y-%m-%d")

    return df

def fixdate_matl(df,cols):
    try:
        for col in cols :
            df[col] = pd.TimedeltaIndex(abs(df[col]),unit='d') + datetime.datetime(1899, 12, 30)
    except : 
        pass

    return df

def calculate_dn(P, T, type):
    if type == "SW":
        factor1 = 1 if P == "SW" else 0 if P == "FW" else None
    elif type == "FW":
        factor1 = 0 if P == "SW" else 1 if P == "FW" else None

    factor2 = {
        15: 0.5, 20: 0.75, 25: 1, 32: 1.25, 
        40: 1.5, 50: 2, 65: 2.5, 80: 3
    }.get(T, T / 25 if T >= 100 else None)

    return factor1 * factor2 if factor1 is not None and factor2 is not None else None

if __name__ == '__main__':
    allcols  =  ['LOI Date',
                'KOM Date',
                'Schedule',
                'Welding Map Date',
                'Receive Dwg',
                'Start Fabric',
                'NS',
                'AREA',
                'System',
                'Priority',
                'Sub Area',
                'Drawing No.',
                'LINE NO',
                'LINE CODE',
                "MAT'L LINE",
                "MAT'L CODE",
                'PAGE',
                'REV',
                'FW / SW',
                'JOINT TYPE',
                'JOINT NO',
                'PIPE SPOOL',
                'DN',
                'DIA-INCH PLAN SW',
                'DIA-INCH PLAN FW',
                'TOTAL',
                'Comment',
                'P LEGEND',
                'P MATL CODE',
                'P TYPE',
                'P MATL TYPE',
                'P GROUP MATL',
                'P CLASS / SCH',
                'P BASE SIZE 1',
                'P BASE SIZE 2',
                'P THK 1',
                'P Standard Matl Name',
                'P QTY MATL',
                'P UOM',
                'S MATL CODE',
                'S TYPE',
                'S MATL TYPE',
                'S GROUP MATL',
                'S CLASS / SCH',
                'S BASE SIZE 1',
                'S BASE SIZE 2',
                'S THK 1',
                'S Standard Matl Name',
                'S QTY MATL',
                'S UOM',
                'T MATL CODE',
                'T TYPE',
                'T MATL TYPE',
                'T GROUP MATL',
                'T CLASS / SCH',
                'T BASE SIZE 1',
                'T BASE SIZE 2',
                'T THK 1',
                'T Standard Matl Name',
                'T QTY MATL',
                'T UOM',
                'MATL REMARKS',
                'P SR / MIR NO',
                'P SR DATE',
                'P PIC',
                'P QTY TOTAL',
                'P QTY HAULING',
                'P QTY OUTSTANDING',
                'P QTY MAPPING',
                'P MAPPING BALANCE',
                'S SR / MIR NO',
                'S SR DATE',
                'S PIC',
                'S QTY TOTAL',
                'S QTY HAULING',
                'S QTY OUTSTANDING',
                'S QTY MAPPING',
                'S MAPPING BALANCE',
                'T SR / MIR NO',
                'T SR DATE',
                'T PIC',
                'T QTY TOTAL',
                'T QTY HAULING',
                'T QTY OUTSTANDING',
                'T QTY MAPPING',
                'T MAPPING BALANCE',
                "MIR REMARKS",
                'FIT-UP RECORD SW',
                'FIT-UP RECORD FW',
                'FIT-UP RECORD DATE',
                'FU SPV',
                'WELDING RECORD SW',
                'WELDING RECORD FW',
                'WELDING RECORD DATE',
                'W SPV',
                'W STAMP',
                'QAQC AFI F/U DATE',
                'QAQC AFI F/U NO',
                'QAQC AFI F/U RES',
                'QAQC VISUAL DATE',
                'QAQC VISUAL NO',
                'QAQC VISUAL RES',
                'PPC CLAIM REPORT FABS 30%',
                'PPC CLAIM REPORT INSTALL 60%',
                'PPC CLAIM REPORT PUNCHLIST 10%',
                'PPC CLAIM REPORT TOTAL',
                'SUMMARY AFI',
                'LAST PERIOD',
                'THIS PERIOD',
                'CUMM',
                'Remarks']
    
    mtocols =  ["No.", 
                "LINE NO", 
                "Dwg No.",
                "Dwg Rev", 
                "PID No", 
                "PID Rev", 
                "Page", 
                "Detail", 
                "Code", 
                "Description", 
                "TYPE", 
                "MATL TYPE", 
                "GROUP MATL", 
                "BASE SIZE 1", 
                "BASE SIZE 2", 
                "CLASS / SCH", 
                "THK 1", 
                "Standard Matl Name", 
                "MATL CODE", 
                "UOM", 
                "QTY", 
                "Remark"
                ]

    mirmastercols = ["No Item",
                    "Project/Vendor Code",
                     "NS",
                     "SR / MIR NO",
                     "SR DATE",
                     "Date Action",
                     "PIC",
                     "TYPE",
                     "MATL TYPE",
                     "GROUP MATL",
                     "BASE SIZE 1",
                     "BASE SIZE 2",
                     "CLASS / SCH",
                     "THK 1",
                     "Standard Matl Name",
                     "MATL CODE",
                     "Ident Code",
                     "Tag Number / PO Number",
                     "QTY TOTAL",
                     "QTY HAULING",
                     "QTY OUTSTANDING",
                     "UOM",
                     "Status Hauling",
                     "Status SR",
                     "Remarks 1",
                     "Remarks 2",
                     "SML",
                     "SMOA"
                     ]     
    
    datecols = ['LOI Date', 
                'KOM Date', 
                'Welding Map Date', 
                "Receive Dwg", 
                'Start Fabric', 
                'Schedule', 
                'FIT-UP RECORD DATE', 
                'WELDING RECORD DATE',
                'QAQC AFI F/U DATE',
                'QAQC VISUAL DATE',
                'P SR DATE',
                'S SR DATE',
                'T SR DATE',
                ]
    
    dbexcols = ['MATL CODE',
                'TYPE',
                'MATL TYPE',
                'GROUP MATL',
                'CLASS / SCH',
                'BASE SIZE 1',
                'BASE SIZE 2',
                'THK 1',
                'Standard Matl Name',
                'UOM',]

    for i in range (1, 11):
        mirmastercols.append(f"HAULING QTY {i}")
        mirmastercols.append(f"HAULING DATE {i}")
        mirmastercols.append(f"HAULING PIC {i}")
    engcols = allcols[:allcols.index("Comment")+1]
    constructioncols = allcols[allcols.index("FIT-UP RECORD SW"):allcols.index("W STAMP")+1]
    engallcols = engcols + constructioncols
    qccols = allcols[allcols.index("QAQC AFI F/U DATE") : allcols.index("QAQC VISUAL RES")+1]
    qcallcols = engallcols + qccols
    ppcols = allcols[allcols.index("PPC CLAIM REPORT FABS 30%"):]
    ppcallcols = qcallcols + ppcols
    progressallcols = engallcols+qccols+ppcols
    matlallcols = allcols[:allcols.index("MATL REMARKS")+1]
    matlcols = allcols[allcols.index('P LEGEND'):allcols.index('MATL REMARKS')+1]
    srallcols = allcols[:allcols.index("MIR REMARKS")+1]
    srcols = allcols[allcols.index("P SR / MIR NO"):allcols.index("MIR REMARKS")+1]
    fullmatlcols = engcols + matlcols + srcols
    mh_progress = allcols[:allcols.index("P UOM")+1] + \
                  allcols[allcols.index('P SR / MIR NO'):allcols.index('P MAPPING BALANCE')+1] + \
                  allcols[allcols.index('S MATL CODE'):allcols.index('S UOM')+1] + \
                  allcols[allcols.index('S SR / MIR NO'):allcols.index('S MAPPING BALANCE')+1] + \
                  allcols[allcols.index('T MATL CODE'):allcols.index('MATL REMARKS')+1] + \
                  allcols[allcols.index('T SR / MIR NO'):allcols.index('MIR REMARKS')+1]

    normal_datecols = datecols[:datecols.index('QAQC VISUAL DATE')+1]
    matl_datecols = datecols[datecols.index("P SR DATE"):datecols.index("T SR DATE")+1]
    first_datecols = datecols[:datecols.index('Schedule')+1]

    engineerdata_path = "ENGINEER_PLACEHOLDER"
    ppc_path = "PPC_PLACEHOLDER"
    qaqc_path = "QAQC_PLACEHOLDER"

    mto_path = "MTO_PLACEHOLDER"
    mir_path = "MIR_PLACEHOLDER"

    masterdata_path = "MASTERDATA_PLACEHOLDER/masterdata.xlsx"
    dashboard_path = "DASHBOARD_PLACEHOLDER"

    backupmaster_path = "MASTERDATA_PLACEHOLDER/BACKUP"

    sheet_name = "Monitoring Piping"
    sheet_name2 = "Updates from Engineer"
    sheet_name3 = "Conflict"
    sheet_name4 = "DB MATL"
    sheet_name5 = "DB Matl"

    try:
        logging.basicConfig(
        level=logging.INFO,  # Level logging hanya untuk info utama
        format='%(asctime)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler("pipeline_log_main.txt"),  # Log ke file dengan ekstensi .txt
            logging.StreamHandler()  # Log ke konsol
            ]
        )
        warnings.filterwarnings("ignore")

        logging.info("============ Starting the Program ============")

        logging.info("Load Data")
        engdata_all = load_data(engineerdata_path, sheet_name, 3, engallcols, True)
        engdata_all['DIA-INCH PLAN SW'] = engdata_all.apply(lambda row: calculate_dn(row['FW / SW'], row['DN'], "SW"), axis=1)
        engdata_all['DIA-INCH PLAN FW'] = engdata_all.apply(lambda row: calculate_dn(row['FW / SW'], row['DN'], "FW"), axis=1)
        engdata_all["TOTAL"] = engdata_all['DIA-INCH PLAN SW'] + engdata_all['DIA-INCH PLAN FW']
        engdata_all["FIT-UP RECORD SW"] = engdata_all.apply(lambda x: x["DIA-INCH PLAN SW"] if pd.notna(x["FIT-UP RECORD DATE"]) else np.nan, axis=1)
        engdata_all["FIT-UP RECORD FW"] = engdata_all.apply(lambda x: x["DIA-INCH PLAN FW"] if pd.notna(x["FIT-UP RECORD DATE"]) else np.nan, axis=1)
        engdata_all["WELDING RECORD SW"] = engdata_all.apply(lambda x: x["DIA-INCH PLAN SW"] if pd.notna(x["WELDING RECORD DATE"]) else np.nan, axis=1)
        engdata_all["WELDING RECORD FW"] = engdata_all.apply(lambda x: x["DIA-INCH PLAN FW"] if pd.notna(x["WELDING RECORD DATE"]) else np.nan, axis=1)
        engdata_joint = engdata_all[engcols + ["PK"]]
        engdata_conflict = load_data(engineerdata_path, sheet_name3, 3, engallcols, True)
        ppc_all = load_data(ppc_path, sheet_name, 3, ppcallcols, True)
        ppc_all['DIA-INCH PLAN SW'] = ppc_all.apply(lambda row: calculate_dn(row['FW / SW'], row['DN'], "SW"), axis=1)
        ppc_all['DIA-INCH PLAN FW'] = ppc_all.apply(lambda row: calculate_dn(row['FW / SW'], row['DN'], "FW"), axis=1)
        ppc_all["TOTAL"] = ppc_all['DIA-INCH PLAN SW'] + ppc_all['DIA-INCH PLAN FW']
        ppc_all["FIT-UP RECORD SW"] = ppc_all.apply(lambda x: x["DIA-INCH PLAN SW"] if pd.notna(x["FIT-UP RECORD DATE"]) else np.nan, axis=1)
        ppc_all["FIT-UP RECORD FW"] = ppc_all.apply(lambda x: x["DIA-INCH PLAN FW"] if pd.notna(x["FIT-UP RECORD DATE"]) else np.nan, axis=1)
        ppc_all["WELDING RECORD SW"] = ppc_all.apply(lambda x: x["DIA-INCH PLAN SW"] if pd.notna(x["WELDING RECORD DATE"]) else np.nan, axis=1)
        ppc_all["WELDING RECORD FW"] = ppc_all.apply(lambda x: x["DIA-INCH PLAN FW"] if pd.notna(x["WELDING RECORD DATE"]) else np.nan, axis=1)
        ppc_all["PPC CLAIM REPORT FABS 30%"] = ppc_all.apply(lambda x: x["TOTAL"]*0.3 if pd.notna(x["QAQC VISUAL DATE"]) else np.nan, axis=1)
        ppc_all["PPC CLAIM REPORT INSTALL 60%"] = ppc_all.apply(lambda x: x["TOTAL"]*0.6 if pd.notna(x["QAQC VISUAL DATE"]) else np.nan, axis=1)
        ppc_all["PPC CLAIM REPORT PUNCHLIST 10%"] = ppc_all.apply(lambda x: x["TOTAL"]*0.1 if pd.notna(x["QAQC VISUAL DATE"]) else np.nan, axis=1)
        ppc_all["PPC CLAIM REPORT TOTAL"] = ppc_all["PPC CLAIM REPORT FABS 30%"] + ppc_all["PPC CLAIM REPORT INSTALL 60%"] + ppc_all["PPC CLAIM REPORT PUNCHLIST 10%"]
        ppc_all["CUMM"] = ppc_all["PPC CLAIM REPORT TOTAL"]
        ppc_cons = ppc_all[constructioncols + ["PK"]]
        ppc_claim = ppc_all[ppcols + ["PK"]]
        ppc_conflict = ppc_all[engallcols + ["PK"]]
        ppc_without_afi = ppc_all[engallcols + ppcols + ["PK"]]
        qc_all = load_data(qaqc_path, sheet_name, 3, qcallcols, True)
        qc_afi = qc_all[qccols + ["PK"]]
        masterdata = load_data(masterdata_path, sheet_name, 0, None)
        masterdata_updates = load_data(masterdata_path, sheet_name2, 0, None)
        dbmatl_mto = load_data(mto_path, sheet_name4,0, dbexcols)
        dbmatl_mir = load_data(mir_path, sheet_name5, 0,None)
        matldata_pure = load_data(mto_path, sheet_name, 3, matlallcols, True)
        srdata_all = load_data(mir_path, sheet_name, 3, srallcols, True)
        srdata = srdata_all[srcols + ["PK"]]
        mtodata = load_data(mto_path, "MTO", 3, mtocols)
        mirdata = load_data(mir_path, "MASTER", 2, mirmastercols)

        logging.info("Standardization Data Types")
        engdata_all["PK"] = engdata_all["PK"].astype(str)
        engdata_joint["PK"] = engdata_joint["PK"].astype(str)
        engdata_conflict["PK"] = engdata_conflict["PK"].astype(str)
        ppc_all["PK"] = ppc_all["PK"].astype(str)
        ppc_cons["PK"] = ppc_cons["PK"].astype(str)
        ppc_claim["PK"] = ppc_claim["PK"].astype(str)
        qc_all["PK"] = qc_all["PK"].astype(str)
        qc_afi["PK"] = qc_afi["PK"].astype(str)
        masterdata["PK"] = masterdata["PK"].astype(str)
        dbmatl_mto['MATL CODE'] = dbmatl_mto[["TYPE","MATL TYPE","BASE SIZE 1","BASE SIZE 2","CLASS / SCH","GROUP MATL","THK 1"]].apply(lambda x: ' '.join(x.dropna().astype(str)), axis=1)
        dbmatl_mir['MATL CODE'] = dbmatl_mir[["TYPE","MATL TYPE","BASE SIZE 1","BASE SIZE 2","CLASS / SCH","GROUP MATL","THK 1"]].apply(lambda x: ' '.join(x.dropna().astype(str)), axis=1)
        matldata_all = pd.merge(matldata_pure, dbmatl_mto, left_on="P MATL CODE", right_on="MATL CODE", how="left")
        matldata_all = pd.merge(matldata_all, dbmatl_mto, left_on="S MATL CODE", right_on="MATL CODE", how="left", suffixes=("","_S"))
        matldata_all = pd.merge(matldata_all, dbmatl_mto, left_on=matldata_all["T MATL CODE"].astype(str), right_on="MATL CODE", how="left", suffixes=("","_T"))
        mirdata["TEMP_PK"] = mirdata[["SR / MIR NO", "MATL CODE"]].apply(lambda x: ' '.join(x.dropna().astype(str)), axis=1)
        mirdata["QTY HAULING"] = mirdata[[f"HAULING QTY {i}" for i in range(1, 11)]].sum(axis=1)
        mirdata["QTY OUTSTANDING"] = mirdata["QTY TOTAL"] - mirdata["QTY HAULING"]
        mirdata = pd.merge(mirdata, dbmatl_mir, left_on="MATL CODE", right_on="MATL CODE", how = 'left', suffixes=("", "_dup"))
        mtodata = pd.merge(mtodata, dbmatl_mto, left_on="MATL CODE", right_on="MATL CODE", how = 'left', suffixes=("", "_dup")) 
        for i in dbexcols :
            matldata_all["P "+ i] = matldata_all["P "+ i].fillna(matldata_all[i])
            matldata_all["S "+ i] = matldata_all["S "+ i].fillna(matldata_all[i+"_S"])
            matldata_all["T "+ i] = matldata_all["T "+ i].fillna(matldata_all[i+"_T"])
            matldata_all.drop(columns = i, inplace=True)
            matldata_all.drop(columns = i+"_S", inplace=True)
            matldata_all.drop(columns = i+"_T", inplace=True)
            if i != "MATL CODE":
                mirdata[i] = mirdata[i].fillna(mirdata[i+"_dup"])
                mirdata.drop(columns = i+"_dup", inplace = True)
                mtodata[i] = mtodata[i].fillna(mtodata[i+"_dup"])
                mtodata.drop(columns = i+"_dup", inplace = True)
        matldata = matldata_all[matlcols+["PK"]]
        matldata_all["PK"] = matldata_all["PK"].astype(str)
        matldata["PK"] = matldata["PK"].astype(str)
        mirdata["Status Hauling"] = mirdata["QTY OUTSTANDING"].fillna(0).apply(lambda x: "OPEN" if x != 0 else "CLOSE")
        srdata_all["PK"] = srdata_all["PK"].astype(str)
        srdata["PK"] = srdata["PK"].astype(str)
        srdata_all["P TEMP_PK"] = srdata_all[["P SR / MIR NO", "P MATL CODE"]].apply(lambda x: ' '.join(x.dropna().astype(str)), axis=1)
        srdata_all["S TEMP_PK"] = srdata_all[["S SR / MIR NO", "S MATL CODE"]].apply(lambda x: ' '.join(x.dropna().astype(str)), axis=1)
        srdata_all["T TEMP_PK"] = srdata_all[["T SR / MIR NO", "T MATL CODE"]].apply(lambda x: ' '.join(x.dropna().astype(str)), axis=1)

        sralldata_merge = pd.merge(srdata_all, mirdata, left_on= "P TEMP_PK", right_on="TEMP_PK",  how = "left", suffixes=("","_P"))
        sralldata_merge = pd.merge(sralldata_merge, mirdata, left_on= "S TEMP_PK", right_on="TEMP_PK", how = "left", suffixes=("","_S"))
        sralldata_merged = pd.merge(sralldata_merge, mirdata, left_on= "T TEMP_PK", right_on="TEMP_PK", how = "left", suffixes=("","_T"))


        for i in ["SR DATE", "PIC", "QTY TOTAL", "QTY HAULING"]:
            sralldata_merged["P " + i] = sralldata_merged["P " + i].fillna(sralldata_merged[i])
            sralldata_merged["S " + i] = sralldata_merged["S " + i].fillna(sralldata_merged[i+"_S"])
            sralldata_merged["T " + i] = sralldata_merged["T " + i].fillna(sralldata_merged[i+"_T"])
        sralldata_merged.drop(columns=sralldata_merged.columns[sralldata_merged.columns.get_loc("TEMP_PK"):], inplace=True)
        srdata = sralldata_merged[srcols + ["PK"]]
        mirdata.drop(columns = "TEMP_PK", inplace=True)


        logging.info("Backup Data")
        backup(engineerdata_path, backupmaster_path, slicing(engineerdata_path))
        backup(ppc_path, backupmaster_path, slicing(ppc_path))
        backup(qaqc_path, backupmaster_path, slicing(qaqc_path))
        backup(mto_path, backupmaster_path, slicing(mto_path))
        backup(mir_path, backupmaster_path, slicing(mir_path))

        logging.info("Process Data")
        updates = update_data(engdata_joint, masterdata, "PK")
        md_cons_data = syncronized(engdata_joint, ppc_cons, "PK", "PK", "left")
        md_afi_data = syncronized(md_cons_data,qc_afi, "PK", "PK", "left")
        md_claim_data = syncronized(md_afi_data,ppc_claim, "PK", "PK", "left")
        md = md_claim_data[ppcallcols + ["PK"]]
        md_for_ppc = pd.merge(ppc_all, md_afi_data, on = "PK", how = "outer", suffixes=("","_dup"))
        for col in ppc_all:
            if col in md_afi_data.columns and col !="PK":
                md_for_ppc[col] = md_for_ppc[col].fillna(md_for_ppc[col+"_dup"])
                md_for_ppc.drop(columns = [col+"_dup"], inplace=True)
        md_for_ppc = md_for_ppc[ppcallcols]
        md_for_qc = md_afi_data[qcallcols]
        md_mto = syncronized(engdata_joint, matldata, "PK", "PK", "left")
        md_sr = syncronized(md_mto,srdata, "PK", "PK", "left")
        md_for_mto = md_mto[matlallcols]
        md_for_sr = md_sr[mh_progress]
        md = pd.merge(md,md_sr[matlcols + srcols + ["PK"]], on = "PK", how = "outer", suffixes=("","_dup"))
        md = md[allcols+["PK"]]
        conflict = pd.merge(ppc_conflict, pd.DataFrame(engdata_joint["PK"]), on = "PK", how = 'outer', indicator = True)
        conflict = conflict[conflict["_merge"] == 'left_only']
        conflict = conflict[engallcols + ["PK"]]
        still_conflict = pd.merge(pd.DataFrame(engdata_conflict["PK"]), conflict, on = "PK",how = "outer", indicator=True)
        still_conflict = still_conflict[still_conflict["_merge"] == "right_only"]
        still_conflict = still_conflict[engallcols]
        engqc = syncronized(engdata_all, md_afi_data[qccols+["PK"]], "PK", "PK", 'left')
        engqc = syncronized(engqc, md_claim_data[ppcols+["PK"]], "PK", "PK", 'left')
        engqc = syncronized(engqc, matldata, "PK", "PK", 'left')
        engqc = syncronized(engqc, srdata, "PK", "PK", 'left')

        logging.info("Fixing Datetime Format")
        md = fixdate(md, datecols)
        md_for_qc = fixdate(md_for_qc, normal_datecols)
        md_for_ppc = fixdate(md_for_ppc, normal_datecols)
        md_for_mto = fixdate_matl(md_for_mto, first_datecols)
        md_for_sr = fixdate_matl(md_for_sr, matl_datecols)

        logging.info("Writing Data")
        write_data(md, masterdata_path, sheet_name, 2, 1, "PK", "claim")
        write_data(md_for_qc, qaqc_path, sheet_name, 5, 1, typedf="progress")
        write_data(md_for_ppc, ppc_path, sheet_name, 5, 1, typedf="claim")
        write_data(md_for_mto, mto_path, sheet_name, 5, 1, typedf="matl")
        write_data(md_for_sr, mir_path, sheet_name, 5, 1, typedf="mir")
        write_data(still_conflict, engineerdata_path, sheet_name3, 5, 1)
        write_data(updates, masterdata_path, sheet_name2, 2,1)
        write_data(updates, ppc_path, sheet_name2, 2,1)
        write_data(updates, qaqc_path, sheet_name2, 2,1)
        write_data(updates, mto_path, sheet_name2, 2,1)
        write_data(updates, mir_path, sheet_name2, 2,1)

        logging.info("Produce Data")
        production(md, dashboard_path, "Ready for dashboard", "PK", "PK", False)
        production(mtodata, dashboard_path, "MTO for dashboard", instance=True)
        production(mirdata, dashboard_path, "MIR for dashboard", instance=True)
        production(engqc, dashboard_path, "Engineer for dashboard", "PK", "JOINT NO")

        logging.info("Aman bro no error")
        logging.info("TIMAS SUPLINDO - Developed by Auvi A")

    except PermissionError as epe:
        error_message = "Close destination file and try again"
        logging.error(f"An error occurred: {epe}")
        logging.error(error_message)

        for i in range(120, 0, -1):
            print(f"exit in {i}s...", end="\r")
            time.sleep(1)
        pass

    except Exception as e:
        logging.critical("An critical occurred: {e}")
        logging.critical(traceback.format_exc())

        for i in range(120, 0, -1):
            print(f"exit in {i}s...", end="\r")
            time.sleep(1)

    else:
        for i in range(5, 0, -1):
            print(f"exit in {i}s...", end="\r")
            time.sleep(1)

#EoL